# -*- coding: utf-8 -*-
"""
One-time sync script to push all 2,525 existing rows from placements.xlsx to Google Sheets.
"""

import os
import json
import urllib.request
from openpyxl import load_workbook

WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbzak0HaxgSXI-QF2k0xUSIj3gjVIFOtIJ5dJU2hTOLkopmwzO2tKTeLGhlj8SyoqLCa/exec"
EXCEL_PATH = r"c:\Users\cheth\OneDrive\Desktop\telebot\placements.xlsx"
LAST_ID_PATH = r"c:\Users\cheth\OneDrive\Desktop\telebot\last_id.txt"

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} not found.")
        return

    print("Loading placements.xlsx...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Read last_id from last_id.txt
    last_id = 20749
    if os.path.exists(LAST_ID_PATH):
        try:
            with open(LAST_ID_PATH, 'r', encoding='utf-8') as f:
                last_id = int(f.read().strip())
        except ValueError:
            pass

    keys = [
        'date', 'company', 'role', 'category', 'salary_probation', 'salary_post_probation',
        'salary_raw', 'location', 'eligibility', 'registration_link', 'deadline',
        'needs_review', 'message_link', 'raw_message'
    ]

    all_rows = []
    for row in range(2, ws.max_row + 1):
        row_dict = {}
        for col_idx, k in enumerate(keys, 1):
            val = ws.cell(row=row, column=col_idx).value
            row_dict[k] = str(val) if val is not None else ""
        
        # Categorize
        role = row_dict['role'].lower()
        comp = row_dict['company'].lower()
        raw = row_dict['raw_message'].lower()
        combined = f"{role} {comp} {raw}"

        if any(kw in role for kw in ['sales', 'bda', 'bde', 'business development', 'client handling', 'lead generation', 'inside sales', 'telecaller', 'prospecting', 'account executive']) or any(kw in combined for kw in ['business development executive', 'business development associate', 'bde role', 'bda role']):
            row_dict['category'] = 'Sales & Business Dev'
        elif any(kw in role for kw in ['marketing', 'content writer', 'seo', 'social media', 'digital marketing', 'brand', 'media', 'growth hacker', 'copywriter']) or any(kw in combined for kw in ['marketing intern', 'digital marketing executive']):
            row_dict['category'] = 'Marketing'
        elif any(kw in role for kw in ['software', 'sde', 'developer', 'full stack', 'frontend', 'backend', 'devops', 'qa', 'testing', 'data analyst', 'data scientist', 'system engineer', 'trainee engineer', 'programmer', 'cyber', 'cloud', 'ui/ux', 'web', 'mobile', 'ai', 'ml', 'code', 'python', 'java', 'c++', 'tech', 'genc', 'analyst trainee', 'internship', 'consultant']) or any(kw in combined for kw in ['software development engineer', 'programmer analyst', 'system engineer', 'genc', 'hackwithinfy', 'codevita']):
            row_dict['category'] = 'IT & Software'
        else:
            row_dict['category'] = 'Core & Other'

        all_rows.append(row_dict)

    print(f"Read {len(all_rows)} rows from Excel.")

    # Push to Google Sheets in batches of 100 rows
    batch_size = 100
    for i in range(0, len(all_rows), batch_size):
        batch = all_rows[i:i + batch_size]
        payload = json.dumps({
            "rows": batch,
            "last_id": last_id
        }).encode('utf-8')

        req = urllib.request.Request(
            WEBHOOK_URL,
            data=payload,
            headers={"Content-Type": "application/json"}
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                res = json.loads(resp.read().decode('utf-8'))
                print(f"Synced rows {i+1} to {min(i+batch_size, len(all_rows))} -> Google Sheets response: {res.get('status')}")
        except Exception as e:
            print(f"Failed sync batch {i+1}: {e}")

    print("\nAll existing placement rows synced to Google Sheets!")

if __name__ == "__main__":
    main()
