# -*- coding: utf-8 -*-
"""
Telegram to Excel Scraper for Placement Posts

This script scrapes job postings from a Telegram group and outputs them into a structured Excel workbook.
"""

import re
import os
import asyncio
from datetime import datetime
from telethon import TelegramClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# Configuration Block
# ==========================================
API_ID   = int(os.environ.get("TELEGRAM_API_ID", 39806525))
API_HASH = os.environ.get("TELEGRAM_API_HASH", "20561160a2f41ad9cbb3f9e45e9bdf67")
GROUP    = os.environ.get("TELEGRAM_GROUP_ID", "-1002020152383")
GOOGLE_WEBHOOK_URL = os.environ.get("GOOGLE_WEBHOOK_URL", "https://script.google.com/macros/s/AKfycbzak0HaxgSXI-QF2k0xUSIj3gjVIFOtIJ5dJU2hTOLkopmwzO2tKTeLGhlj8SyoqLCa/exec")
EXCEL    = 'placements.xlsx'
LAST_ID  = 'last_id.txt'

def sync_to_google_sheet(rows, max_id):
    if not GOOGLE_WEBHOOK_URL or not rows:
        return
    try:
        import urllib.request
        import json
        print("Syncing new rows to Google Sheets...")
        payload = json.dumps({
            "rows": rows,
            "last_id": max_id
        }).encode('utf-8')
        req = urllib.request.Request(
            GOOGLE_WEBHOOK_URL,
            data=payload,
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            print("Successfully synced to Google Sheets!")
    except Exception as e:
        print(f"Google Sheets sync warning: {e}")

# ==========================================
# 1. Last ID Tracking
# ==========================================
def read_last_id():
    if os.path.exists(LAST_ID):
        try:
            with open(LAST_ID, 'r', encoding='utf-8') as f:
                return int(f.read().strip())
        except ValueError:
            pass
    return 0

def write_last_id(msg_id):
    with open(LAST_ID, 'w', encoding='utf-8') as f:
        f.write(str(msg_id))

# ==========================================
# 2. Message Fetcher
# ==========================================
async def fetch_new_messages(client, last_id):
    messages = []
    group_entity = int(GROUP) if str(GROUP).lstrip('-').isdigit() else GROUP
    async for msg in client.iter_messages(group_entity, min_id=last_id, reverse=True):
        if msg.text and not msg.action:
            messages.append(msg)
    return messages

# ==========================================
# 3. Placement Post Filter
# ==========================================
def is_placement_post(msg):
    text = msg.text or ""
    if len(text) < 80:
        return False
    if msg.reply_to is not None:
        return False
    
    keywords = [
        'company', 'role', 'salary', 'compensation', 'ctc', 'lpa', 
        'campus drive', 'job title', 'eligibility', 'register', 
        'recruitment', 'openings', 'hiring', 'walk-in', 'package', 
        'stipend', 'job description', 'vacancy', 'fresher'
    ]
    text_lower = text.lower()
    hits = sum(1 for kw in keywords if kw in text_lower)
    return hits >= 2

# ==========================================
# 4. Parsers
# ==========================================
def extract_company(text):
    """Extract company name using labeled patterns, then careful heuristics."""
    
    # Blacklist of invalid company names / noisy section titles
    invalid_companies = {
        'about company', 'details', 'key details', 'hrms', 'workshop highlights',
        'alert', 'opportunity with', 'remainder', 'registration', 'who can',
        'ask how', 'updates', 'policy', 'note', 'important note', 'final reminder',
        'gentle reminder', 'call for', 'incorrect', 'who can participate',
        'registration deadline', 'test configuration', 'public keys', 'the company',
        'for women coders', 'challenge alert', 'this is an', 'office', 'webinar',
        'you can view', 'we are excited'
    }

    def clean_comp_name(name):
        if not name:
            return None
        # Remove common noise phrases from start
        name = re.sub(r'^(?:Greetings?\s+from|Hurry\s+Up\s*[\-\:]?|Alert\s*[\-\:]?|Attention\s+to\s+All\s+(?:Registered\s+)?Students\s+(?:for\s+)?|Important\s+Note\s+(?:for\s+[\w\s]+\s+on\s+)?)\s*', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'\s*-\s*Pune\'s\s+Top.*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'\s*[\-\:]\s*All\s+Branches.*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'[^\w\s\-\&\.\,\']', '', name).strip('- ').strip()
        
        if name.lower() in invalid_companies or any(inv in name.lower() for inv in ['registration', 'updates', 'reminder', 'highlight', 'webinar', 'volunteer']):
            return None
        if len(name.split()) > 7 or len(name) > 60:
            return None
        return name

    # Priority 1: Explicit "Company: X" label
    m1 = re.search(r'Company\s*:\s*(.+)', text, re.IGNORECASE)
    if m1:
        val = clean_comp_name(m1.group(1).strip().split('\n')[0].strip())
        if val: return val
    
    # Priority 2: "Greetings from X", "from X!", "by X"
    m2 = re.search(r'(?:Greetings\s+from|from|by)\s+([A-Z][A-Za-z0-9\s\-\&\.]{2,30}?)(?:[\.!\n]|\s+is\s+|\s+has\s+|\s+invites|\s+hiring|$)', text)
    if m2:
        val = clean_comp_name(m2.group(1).strip())
        if val and len(val.split()) <= 5: return val
    
    # Priority 3: "X is hiring" / "X is currently hiring" / "X wants to hire"
    m3 = re.search(r'^([A-Z][A-Za-z0-9\s\-\&]{2,30}?)\s+(?:is\s+(?:currently\s+)?(?:hiring|recruiting)|wants to hire)', text, re.MULTILINE)
    if m3:
        val = clean_comp_name(m3.group(1).strip())
        if val: return val

    # Priority 4: "X Campus Drive" or "X Off-Campus" or "X Internship"
    m4 = re.search(r'([A-Z][A-Za-z0-9\s\-\&]{2,30}?)\s+(?:Campus\s+(?:Drive|Recruitment)|Off[- ]?Campus|Internship\s+Hiring|Innovation\s+Challenge|CodeVita|Hackathon)', text)
    if m4:
        val = clean_comp_name(m4.group(1).strip())
        if val: return val

    # Priority 5: Known Corporate / Tech Companies List
    known_companies = [
        'Accenture', 'Adobe', 'Airbus', 'Amazon', 'Atlassian', 'Atlas Copco',
        'BTL India', 'Bosch', 'Capgemini', 'Chemiasoft', 'Cognite', 'Cognizant',
        'DRDO', 'DeltaX', 'EdgeVerve', 'Edgeverve Systems', 'EY', 'Finacle',
        'Flipkart', 'Fractal', 'Geeks Kepler', 'Google', 'HCLTECH', 'HCL',
        'IBM', 'Infosys', 'Internshala', 'Kirloskar', 'KNS Properties', "L'Oréal",
        'Luminous Power Technologies', 'Mayura Consultancy Services', 'Microsoft',
        'NEXANOVA PROTECH', 'P&G', 'Proctor and Gamble', 'PeopleHum', 'Recruit CRM',
        'Paxcom India', 'Profound Edutech', 'Reliance Industries', 'RIL',
        'SAP India', 'SAP LABS', 'SAP', 'Schneider Electric', 'ServiceNow',
        'Smarsh', 'SmartFalcon', 'StoneX', 'TCS', 'Tata Steel', 'Tata Elxsi',
        'Talent Battle', 'Talview', 'Tally Solutions', 'Teamlease', 'TheMathCompany',
        'TVS Credit', 'TVS', 'Unisys', 'Unstop', 'VISA', 'Walmart', 'ZS', 'cvDragon'
    ]
    for kc in known_companies:
        if re.search(rf'\b{re.escape(kc)}\b', text, re.IGNORECASE):
            return kc

    # Priority 6: First line ONLY if it looks like a clean company name
    first_line = text.split('\n')[0].strip()
    clean_line = re.sub(r'[^\w\s\-,\&]', '', first_line).strip()
    clean_line = re.sub(r'\s*-\s*\d{4}', '', clean_line).strip()
    suffixes = [r'Campus\s+drive', r'Hiring', r'Recruitment', r'Placements?', r'Drive']
    for suf in suffixes:
        clean_line = re.sub(rf'\b{suf}\b', '', clean_line, flags=re.IGNORECASE).strip()
    clean_line = clean_line.strip('- ').strip()
    
    greetings = ['dear', 'hello', 'hi ', 'attention', 'note', 'important', 'reminder',
                 'final reminder', 'students', 'all ', 'this ', 'today', 'we ', 'the ',
                 'registration', 'who can', 'ask how', 'only for', 'gentle reminder',
                 'call for', 'incorrect', 'job description']
    if clean_line and len(clean_line.split()) <= 4:
        if not any(clean_line.lower().startswith(g) for g in greetings):
            val = clean_comp_name(clean_line)
            if val: return val
    
    return None

def classify_category(row):
    """Categorize row into IT & Software, Sales & Business Dev, Marketing, or Core & Other."""
    role = (row.get('role') or '').lower()
    comp = (row.get('company') or '').lower()
    raw = (row.get('raw_message') or '').lower()
    combined = f"{role} {comp} {raw}"

    # 1. Sales & Business Development
    sales_kws = ['sales', 'bda', 'bde', 'business development', 'client handling', 'lead generation', 'inside sales', 'telecaller', 'prospecting', 'account executive']
    if any(kw in combined for kw in sales_kws):
        return 'Sales & Business Dev'

    # 2. Marketing
    mkt_kws = ['marketing', 'content writer', 'seo', 'social media', 'digital marketing', 'brand', 'media', 'growth hacker', 'copywriter']
    if any(kw in combined for kw in mkt_kws):
        return 'Marketing'

    # 3. IT & Software / Tech
    it_kws = ['software', 'sde', 'developer', 'full stack', 'frontend', 'backend', 'devops', 'qa', 'testing', 'data analyst', 'data scientist', 'system engineer', 'trainee engineer', 'programmer', 'cyber', 'cloud', 'ui/ux', 'web', 'mobile', 'ai', 'ml', 'coding', 'coder', 'python', 'java', 'c++', 'tech', 'genc', 'analyst trainee', 'hackwithinfy', 'codevita', 'codebrewers', 'grid 6.0', 'telisport']
    if any(kw in combined for kw in it_kws):
        return 'IT & Software'

    return 'Core & Other'

def extract_roles(text):
    """Extract job role/title names — NOT responsibilities, instructions, or URLs."""
    roles = []

    def is_valid_role(r):
        if not r:
            return False
        r_str = r.strip()
        # Reject URLs, links, or file paths
        if re.search(r'https?://|www\.|drive\.google\.com|forms\.gle|\.php|\.pdf', r_str, re.IGNORECASE):
            return False
        # Reject sentences / long instructions / responsibilities
        if len(r_str.split()) > 8 or len(r_str) > 70:
            return False
        if any(w in r_str.lower() for w in ['strictly', 'unplaced', 'eligible', 'selection', 'written test', 'interview', 'round', 'responsibility', 'responsibilities', 'criteria', 'note:']):
            return False
        return True

    # Priority 1: Explicit "Job Title:", "Role:", "Position:", "Designation:", "Profile:"
    m1 = re.search(r'(?:Job Title|Position|Profile|Designation|Role Title)\s*:\s*(.+)', text, re.IGNORECASE)
    if m1:
        val = m1.group(1).strip().split('\n')[0].strip()
        if is_valid_role(val):
            roles.append(val)
            return roles
        
    # Priority 2: "Job Description – X" (role name after dash)
    m2 = re.search(r'Job Description\s*[–\-:]\s*(.+)', text, re.IGNORECASE)
    if m2:
        val = m2.group(1).strip().split('\n')[0].strip()
        if is_valid_role(val):
            roles.append(val)
            return roles

    # Priority 3: "Role: X" (explicit label, not "Role Overview" or "Role & Responsibilities")
    for m in re.finditer(r'(?:^|\n)\s*Role\s*:\s*(.+)', text, re.IGNORECASE):
        val = m.group(1).strip().split('\n')[0].strip()
        if is_valid_role(val) and 'overview' not in val.lower():
            roles.append(val)
    if roles:
        return roles[:5]

    # Priority 4: Numbered items like "1. Software Engineer:" or "1) Software Engineer"
    for line in text.split('\n'):
        m = re.match(r'^\s*\d+[\.\)]\s*([A-Za-z0-9\s\-\/\(\)\&]+?)(?:\s*:|\s*–|\s*-|$)', line)
        if m:
            val = m.group(1).strip()
            if is_valid_role(val) and any(kw in val.lower() for kw in ['engineer', 'developer', 'associate', 'executive', 'analyst', 'trainee', 'intern', 'manager', 'lead', 'consultant', 'bda', 'boe', 'sde', 'get']):
                roles.append(val)
    if roles:
        return roles[:5]

    # Priority 5: Bullet points ONLY in a "hiring for" / "openings" / "positions" section
    lines = text.split('\n')
    role_section = False
    for line in lines:
        line_lower = line.lower().strip()
        if re.search(r'(?:hiring|openings?|positions?|vacancies|following roles|open roles)\s*:?\s*$', line_lower):
            role_section = True
            continue
        if role_section:
            if re.search(r'(?:eligibility|criteria|responsibilit|qualification|selection|process|compensation|salary|about|key resp|requirements|skills)', line_lower):
                role_section = False
                continue
            m = re.match(r'^\s*[•\-\*]\s+(.+)', line, re.UNICODE)
            if m:
                candidate = m.group(1).strip(': ')
                if is_valid_role(candidate):
                    roles.append(candidate)
            elif line.strip() and not re.match(r'^\s*[•\-\*]', line, re.UNICODE):
                role_section = False
    if roles:
        return roles[:5]

    # Priority 6: Fallback scan for standard job titles in placement posts
    known_roles = [
        'Software Development Engineer', 'Software Engineer', 'Full Stack Developer',
        'Frontend Developer', 'Backend Developer', 'DevOps Engineer', 'Data Analyst',
        'Data Scientist', 'Business Analyst', 'Business Development Associate',
        'Business Development Executive', 'Marketing Associate', 'Marketing Executive',
        'System Engineer', 'Graduate Engineer Trainee', 'Programmer Analyst Trainee',
        'Programmer Analyst', 'Associate Software Engineer', 'QA Engineer',
        'Quality Assurance Engineer', 'Testing Engineer', 'Cyber Security Intern',
        'Cloud Engineer', 'Sales Executive', 'Customer Success Associate'
    ]
    for kr in known_roles:
        if re.search(rf'\b{re.escape(kr)}\b', text, re.IGNORECASE):
            roles.append(kr)
            break

    return roles

def extract_salary(text):
    result = {'probation': None, 'post_probation': None, 'raw': None}
    
    # Extract raw salary block
    raw_match = re.search(
        r'(?:Compensation(?: Structure)?|Salary|CTC)\s*:\s*(.+?)(?=\n\s*(?:[A-Z][a-z]+|Eligibility|Location|Selection|Key Resp)\s*:|$)',
        text, re.IGNORECASE | re.DOTALL
    )
    if raw_match:
        result['raw'] = raw_match.group(1).strip()[:300]  # cap raw at 300 chars
    else:
        m_sal = re.search(r'(?:Salary|CTC)\s*:\s*(.+)', text, re.IGNORECASE)
        if m_sal:
            result['raw'] = m_sal.group(1).strip()
            
    # Priority 1: Probation vs Post Probation (check FIRST — more specific)
    m_prob = re.search(r'During Probation[^:]*:\s*₹?\s*([\d,]+)\s*(?:per month|p\.?m\.?|/month)?', text, re.IGNORECASE | re.UNICODE)
    m_post = re.search(r'Post Probation[^:]*:\s*₹?\s*([\d,]+)\s*(?:per month|p\.?m\.?|/month)?', text, re.IGNORECASE | re.UNICODE)
    if m_prob and m_post:
        result['probation'] = f'₹{m_prob.group(1).strip()}/month'
        result['post_probation'] = f'₹{m_post.group(1).strip()}/month'
        # Also check for LPA in post-probation line
        m_post_lpa = re.search(r'Post Probation[^\n]*(\d+(?:\.\d+)?)\s*LPA', text, re.IGNORECASE)
        if m_post_lpa:
            result['post_probation'] = f'{m_post_lpa.group(1)} LPA'
        return result
    
    # Priority 2: Flat salary in Salary/CTC line
    flat_match = re.search(r'(?:Salary|CTC)\s*:\s*([^\n]+)', text, re.IGNORECASE)
    if flat_match:
        val = flat_match.group(1).strip()
        result['probation'] = val
        result['post_probation'] = val
        return result
    
    # Fallback: LPA values (annual) — keep separate from monthly ₹ amounts
    if not result['probation'] and not result['post_probation']:
        lpas = re.findall(r'(\d+(?:\.\d+)?)\s*LPA', text, re.IGNORECASE)
        monthly_rs = re.findall(r'₹\s*([\d,]+)\s*(?:per month|p\.?m\.?|/month)', text, re.IGNORECASE | re.UNICODE)
        
        if lpas:
            lpa_vals = sorted(set(float(l) for l in lpas))
            if len(lpa_vals) >= 2:
                result['probation'] = f'{lpa_vals[0]} LPA'
                result['post_probation'] = f'{lpa_vals[-1]} LPA'
            else:
                result['probation'] = f'{lpa_vals[0]} LPA'
                result['post_probation'] = f'{lpa_vals[0]} LPA'
        elif monthly_rs:
            rs_vals = sorted(set(int(r.replace(',', '')) for r in monthly_rs))
            if len(rs_vals) >= 2:
                result['probation'] = f'₹{rs_vals[0]:,}/month'
                result['post_probation'] = f'₹{rs_vals[-1]:,}/month'
            else:
                result['probation'] = f'₹{rs_vals[0]:,}/month'
                result['post_probation'] = f'₹{rs_vals[0]:,}/month'
                
    return result

def extract_locations(text):
    m = re.search(r'(?:Job\s+)?Locations?\s*:\s*(.+)', text, re.IGNORECASE)
    locations = []
    if m:
        raw_loc = m.group(1)
        locations = [l.strip() for l in re.split(r'[,|]', raw_loc) if l.strip()]
    else:
        cities = ['Bangalore', 'Bengaluru', 'Hyderabad', 'Chennai', 'Mumbai', 'Delhi', 'NCR', 'Pune', 'Noida', 'Gurgaon', 'Gurugram', 'Kolkata', 'Ahmedabad', 'Jaipur', 'Kochi', 'Coimbatore', 'Chandigarh', 'Lucknow', 'Indore', 'Remote', 'Work from Home', 'WFH', 'Pan India']
        for city in cities:
            if re.search(rf'\b{city}\b', text, re.IGNORECASE):
                locations.append(city)
                
    # Deduplicate (case-insensitive)
    seen = set()
    dedup = []
    for loc in locations:
        if loc.lower() not in seen:
            seen.add(loc.lower())
            dedup.append(loc)
            
    return ", ".join(dedup) if dedup else None

def extract_eligibility(text):
    # Match 'Eligibility Criteria:' or 'Eligibility:' or just 'Eligibility' on its own line
    m = re.search(r'Eligibility\s*(?:Criteria)?\s*:?\s*\n?(.*?)(?=\n\s*(?:[A-Z][a-zA-Z\s]+:|Required Skills|Compensation|Selection Process|About|Key Resp)|\Z)', text, re.IGNORECASE | re.DOTALL)
    if m:
        content = m.group(1).strip()
        if not content:
            return None
        # Clean up bullet points
        content = re.sub(r'^\s*[•\-\*]\s*', '', content, flags=re.MULTILINE | re.UNICODE)
        content = re.sub(r'\n+', ' | ', content)
        content = content.strip(' |')
        return content if len(content) > 5 else None
    return None

def extract_links(text):
    urls = re.findall(r'https?://[^\s<>"\'\)]+', text, re.IGNORECASE)
    return " | ".join(urls) if urls else None

def extract_deadline(text):
    # Try specific patterns first — 'by' alone is too common
    patterns = [
        r'(?:on or before|last date[:\s]*)\s*(.+?)(?:\.|\n|$)',
        r'(?:register|apply)\s+(?:by|before)\s+(.+?)(?:\.|\n|using|$)',
        r'(?:before|by)\s+(\d{1,2}[:\d]*\s*(?:AM|PM|am|pm)\s+(?:today|tomorrow|\d{1,2}(?:st|nd|rd|th)?\s+\w+(?:\s+\d{4})?))',
        r'(?:deadline|last date)\s*[:\-]\s*(.+?)(?:\.|\n|$)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if len(val) > 5:  # avoid noise like single words
                return val
    return None

# ==========================================
# 5. Main Parser
# ==========================================
def parse_message(msg):
    """Parse a single message into ONE row dict."""
    text = msg.text or ""
    company = extract_company(text)
    roles = extract_roles(text)
    salary = extract_salary(text)
    location = extract_locations(text)
    eligibility = extract_eligibility(text)
    links = extract_links(text)
    deadline = extract_deadline(text)
    
    # Join multiple roles into one cell with ' | ' separator
    role_str = ' | '.join(roles) if roles else None
    
    chat_id = str(msg.chat_id).replace('-100', '') if msg.chat_id else ''
    
    row = {
        'date': msg.date.strftime('%Y-%m-%d') if msg.date else '',
        'company': company,
        'role': role_str,
        'category': '',  # Will be set below
        'salary_probation': salary['probation'],
        'salary_post_probation': salary['post_probation'],
        'salary_raw': salary['raw'],
        'location': location,
        'eligibility': eligibility,
        'registration_link': links,
        'deadline': deadline,
        'needs_review': 'Yes' if not company or not role_str else 'No',
        'message_link': f'https://t.me/c/{chat_id}/{msg.id}' if chat_id else '',
        'raw_message': text[:500]
    }
    row['category'] = classify_category(row)
    return row

# ==========================================
# 6. Excel Handler
# ==========================================
CATEGORY_SHEETS = ['All Placements', 'IT & Software', 'Sales & Business Dev', 'Marketing', 'Core & Other']

def setup_sheet_headers(ws):
    headers = [
        'Date', 'Company', 'Role', 'Category', 'Salary (Probation)', 'Salary (Post-Probation)', 
        'Salary (Raw)', 'Location', 'Eligibility', 'Registration Link', 
        'Deadline', 'Needs Review', 'Message Link', 'Raw Message'
    ]
    ws.append(headers)
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(bottom=Side(style='thin', color='2F5496'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

def ensure_excel(filepath):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        for cat in CATEGORY_SHEETS:
            if cat not in wb.sheetnames:
                ws = wb.create_sheet(title=cat)
                setup_sheet_headers(ws)
        return wb
    
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = CATEGORY_SHEETS[0]
    setup_sheet_headers(ws_main)
    
    for cat in CATEGORY_SHEETS[1:]:
        ws = wb.create_sheet(title=cat)
        setup_sheet_headers(ws)
        
    return wb

def append_rows_categorized(wb, rows):
    count = 0
    keys = [
        'date', 'company', 'role', 'category', 'salary_probation', 'salary_post_probation',
        'salary_raw', 'location', 'eligibility', 'registration_link', 'deadline',
        'needs_review', 'message_link', 'raw_message'
    ]
    light_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    for r in rows:
        cat = r.get('category', 'Core & Other')
        # Always add to 'All Placements' and its specific category sheet
        target_sheets = [wb['All Placements']]
        if cat in wb.sheetnames:
            target_sheets.append(wb[cat])
            
        for ws in target_sheets:
            ws.append([r[k] for k in keys])
            row_num = ws.max_row
            fill = light_fill if row_num % 2 == 0 else white_fill
            for cell in ws[row_num]:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=False)
        count += 1
    return count

def format_all_sheets(wb):
    col_config = {
        'A': {'width': 12, 'align': 'center'},     # Date
        'B': {'width': 22, 'align': 'left'},        # Company
        'C': {'width': 32, 'align': 'left'},        # Role
        'D': {'width': 20, 'align': 'center'},      # Category
        'E': {'width': 20, 'align': 'center'},      # Salary (Probation)
        'F': {'width': 22, 'align': 'center'},      # Salary (Post-Probation)
        'G': {'width': 30, 'align': 'left'},        # Salary (Raw)
        'H': {'width': 28, 'align': 'left'},        # Location
        'I': {'width': 35, 'align': 'left'},        # Eligibility
        'J': {'width': 40, 'align': 'left'},        # Registration Link
        'K': {'width': 22, 'align': 'center'},      # Deadline
        'L': {'width': 14, 'align': 'center'},      # Needs Review
        'M': {'width': 35, 'align': 'left'},        # Message Link
        'N': {'width': 50, 'align': 'left'},        # Raw Message
    }
    
    review_fill = PatternFill(start_color='FCE4E4', end_color='FCE4E4', fill_type='solid')
    review_font = Font(color='C00000', bold=True)

    for ws in wb.worksheets:
        for col_letter, config in col_config.items():
            ws.column_dimensions[col_letter].width = config['width']
            align = Alignment(horizontal=config['align'], vertical='top', wrap_text=(col_letter in ('G', 'I', 'N')))
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                cell.alignment = align
        
        for row in range(2, ws.max_row + 1):
            cell = ws[f'L{row}']
            if cell.value == 'Yes':
                cell.fill = review_fill
                cell.font = review_font

# ==========================================
# 7. Main Flow
# ==========================================
client = TelegramClient('session', API_ID, API_HASH)

async def main():
    await client.start()
    
    last_id = read_last_id()
    print('Fetching messages...')
    messages = await fetch_new_messages(client, last_id)
    
    if not messages:
        print('No new messages.')
        return
        
    print(f'Processing {len(messages)} messages...')
    
    all_rows = []
    skipped = 0
    for msg in messages:
        if not is_placement_post(msg):
            skipped += 1
            continue
        row = parse_message(msg)
        all_rows.append(row)
        
    max_id = max(m.id for m in messages)
    
    if all_rows:
        wb = ensure_excel(EXCEL)
        count = append_rows_categorized(wb, all_rows)
        format_all_sheets(wb)
        try:
            wb.save(EXCEL)
            print(f'Added {count} rows to {EXCEL}')
            write_last_id(max_id)  # Only update after successful save
            sync_to_google_sheet(all_rows, max_id)
            print(f'Processed {len(messages)} messages ({skipped} skipped). Last ID: {max_id}')
        except PermissionError:
            print(f"Error: {EXCEL} is locked. Please close it in Excel and try again.")
            print("Last ID NOT updated — next run will re-process these messages.")
    else:
        print('No placement posts found in new messages.')
        write_last_id(max_id)  # No rows but still advance the pointer
        print(f'Processed {len(messages)} messages ({skipped} skipped). Last ID: {max_id}')
    
    print('Done.')

if __name__ == "__main__":
    with client:
        client.loop.run_until_complete(main())
